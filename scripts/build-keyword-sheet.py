"""One-off: generate keyword-priority-sheet.xlsx in project root."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "keyword-priority-sheet.xlsx"

ROWS = [
    (
        1,
        "décoration chambre tunisie",
        480,
        "9",
        "I",
        "P0",
        "Highest volume + very easy KD",
    ),
    (
        2,
        "décoration murale extérieure",
        260,
        "21",
        "I",
        "P0",
        "Strong volume; harder KD — one strong guide page",
    ),
    (
        3,
        "décoration murale tunisie",
        210,
        "4",
        "I",
        "P0",
        "Core head term + easiest KD in set",
    ),
    (
        4,
        "décoration café",
        210,
        "25",
        "I, T",
        "P1",
        "B2B angle; higher KD",
    ),
    (
        5,
        "décoration du bureau",
        140,
        "16",
        "I",
        "P1",
        "Mid-funnel; fits product use cases",
    ),
    (
        6,
        "tableau décoratif mural tunisie",
        90,
        "n/a",
        "n/a",
        "P1",
        "High product fit — refresh KD in Semrush when possible",
    ),
    (
        7,
        "décoration murale",
        480,
        "33 (SD Ubersuggest)",
        "C",
        "P1",
        "Pillar page; support with specific posts",
    ),
    (
        8,
        "décoration murale salon",
        110,
        "22 (SD)",
        "C",
        "P1",
        "Sub-cluster under décoration murale",
    ),
    (
        9,
        "décoration murale exterieur",
        70,
        "20 (SD)",
        "C",
        "P2",
        "Align with extérieur pillar to avoid thin duplicate pages",
    ),
    (
        10,
        "décoration murale bois",
        50,
        "13 (SD)",
        "—",
        "P2",
        "Only with clear angle (metal vs bois / alternatives)",
    ),
    (
        11,
        "décoration murale salon tendance",
        50,
        "5 (SD)",
        "—",
        "P2",
        "Trend angle",
    ),
    (
        12,
        "décoration murale chambre",
        30,
        "14 (SD)",
        "—",
        "P2",
        "Supports chambre cluster",
    ),
    (
        13,
        "art mural métal / tableau mural métal (cluster)",
        "~10",
        "5–13 (SD)",
        "—",
        "P3",
        "Low TN volume — use in copy, not as sole pillar",
    ),
    (
        14,
        "cadeau mariage tunisie (cluster)",
        "low",
        "n/a",
        "n/a",
        "P2–P3",
        "Conversion content; small volume in Semrush capture",
    ),
]

PUBLISH_ORDER = [
    "1. décoration chambre tunisie",
    "2. décoration murale tunisie",
    "3. décoration murale extérieure",
    "4. décoration du bureau",
    "5. décoration café",
    "6. tableau décoratif mural tunisie",
    "7. décoration murale salon",
    "8. décoration murale (pillar + internal links to 1–7)",
]

LEGEND = [
    "P0 = ship first",
    "P1 = next wave",
    "P2 = after clusters are solid",
    "P3 = supporting phrases in copy, not pillar pages",
    "",
    "Source: Semrush Tunisia + Ubersuggest Tunisia screenshots (Feb–Mar 2026).",
    "KD n/a = refresh metrics in Semrush when possible.",
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Keywords"

    headers = ["Rank", "Keyword", "Volume", "KD %", "Intent (tool)", "Priority", "Notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in ROWS:
        ws.append(list(row))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["G"].width = 48

    ws2 = wb.create_sheet("Publish order")
    ws2.append(["Suggested publish order"])
    ws2["A1"].font = Font(bold=True)
    for line in PUBLISH_ORDER:
        ws2.append([line])
    ws2.column_dimensions["A"].width = 72

    ws3 = wb.create_sheet("Legend")
    for line in LEGEND:
        ws3.append([line])
    ws3.column_dimensions["A"].width = 72

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
